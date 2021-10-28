import csv
import glob
import ntpath
import os

from openpyxl import load_workbook


def get_first_sheet(path):
    return load_workbook(path).active


def read_cells(path, excel_data):
    try:
        sheet = get_first_sheet(path)

        data = {}

        for key, value in excel_data.items():
            data[key] = {}
            for name, cell in value.items():
                data[key][name] = sheet[cell].value

        return data
    except Exception as error:
        raise Exception(error)


def get_rows_from(path, data):
    sheet = get_first_sheet(path)

    rows_data = []

    for row in list(sheet.iter_rows(values_only=True))[data["from_row"]:]:
        row_data = {}

        for key, column in data["columns"].items():
            row_data[key] = row[column]

        rows_data.append(row_data)

    return rows_data


def read_rows(path):
    _, file_extension = os.path.splitext(path)

    if "csv" in file_extension:
        return read_rows_csv(path)

    return read_rows_excel(path)


def read_rows_excel(path):
    sheet = get_first_sheet(path)

    rows = []

    for row in sheet.iter_rows(values_only=True):
        rows.append(row[0])

    return rows


def read_rows_csv(path):
    with open(path, encoding="utf-8") as csv_file:
        reader = csv.reader(csv_file, delimiter=";")

        rows = []

        for row in reader:
            rows.append(row[0])

    return rows


def get_recent_from_directory(directory_path):
    supported_extensions = ["*.xlsx", "*.xlsm", "*.csv"]

    files = []

    for ext in supported_extensions:
        for file in glob.glob(os.path.join(directory_path, ext)):
            files.append(file)

    return sorted(files, key=os.path.getatime)


def find_by_name(directory_path, find_term):
    files = get_recent_from_directory(directory_path)

    for path in files:
        filename = ntpath.basename(path)

        if find_term.lower() in filename.lower():
            return path

    raise Exception(f"Nie znaleziono pliku {find_term}")


def filter_out_none(rows, key):
    target_rows = []

    for row in rows:
        if row[key]:
            target_rows.append(row)

    return target_rows
